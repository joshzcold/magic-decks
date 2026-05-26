#!/usr/bin/env -S uv run --script
#
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "openpyxl"
# ]
# ///
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_csv(csv_path: Path):
    with csv_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return [], []
    header = rows[0]
    data = rows[1:]
    return header, data


def write_sheet(ws, header, rows, start_row=1, start_col=1, freeze_row=1, scale=3):
    for col_index, value in enumerate(header, start=start_col):
        cell = ws.cell(row=start_row, column=col_index, value=value)
        cell.font = Font(bold=True, size=14)
    row_offset = start_row + 1
    for row_index, row in enumerate(rows, start=row_offset):
        for col_index, value in enumerate(row, start=start_col):
            if isinstance(value, str):
                lowered = value.strip().lower()
                if lowered == "true":
                    ws.cell(row=row_index, column=col_index, value=True)
                    continue
                if lowered == "false":
                    ws.cell(row=row_index, column=col_index, value=False)
                    continue
            if isinstance(value, str) and value.startswith("="):
                ws.cell(row=row_index, column=col_index, value=value)
            else:
                ws.cell(row=row_index, column=col_index, value=value)
    if freeze_row:
        ws.freeze_panes = ws.cell(row=freeze_row + 1, column=1)

    max_rows = row_offset + len(rows) - 1
    max_cols = start_col + len(header) - 1
    if max_rows >= start_row + 1:
        for row_index in range(start_row + 1, max_rows + 1):
            ws.row_dimensions[row_index].height = 15 * scale * 3
    image_col_index = None
    for idx, name in enumerate(header, start=start_col):
        if name == "Image":
            image_col_index = idx
            break

    if max_cols >= start_col:
        for col_index in range(start_col, max_cols + 1):
            letter = get_column_letter(col_index)
            if image_col_index == col_index:
                ws.column_dimensions[letter].width = 8.43 * scale * 2
            else:
                ws.column_dimensions[letter].width = 8.43 * 2


def apply_conditional_formatting(ws, header, data_row_start=2):
    if not header:
        return
    max_row = ws.max_row
    header_map = {name: idx + 1 for idx, name in enumerate(header)}

    have_col = header_map.get("Have Physical Copy")
    cut_col = header_map.get("Cut")

    if have_col:
        col_letter = get_column_letter(have_col)
        data_range = f"{col_letter}{data_row_start}:{col_letter}{max_row}"
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f"=OR(${col_letter}{data_row_start}=TRUE,LOWER(TEXT(${col_letter}{data_row_start},\"@\"))=\"true\")"],
                fill=green_fill,
            ),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f"=OR(${col_letter}{data_row_start}=FALSE,LOWER(TEXT(${col_letter}{data_row_start},\"@\"))=\"false\")"],
                fill=yellow_fill,
            ),
        )

    if cut_col:
        col_letter = get_column_letter(cut_col)
        data_range = f"{col_letter}{data_row_start}:{col_letter}{max_row}"
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f"=OR(${col_letter}{data_row_start}=TRUE,LOWER(TEXT(${col_letter}{data_row_start},\"@\"))=\"true\")"],
                fill=red_fill,
            ),
        )


def add_missing_copy_total(ws, header, row_count, data_row_start=2):
    if not header or row_count <= 0:
        return
    header_map = {name: idx + 1 for idx, name in enumerate(header)}
    have_col = header_map.get("Have Physical Copy")
    price_col = header_map.get("Price")
    if not have_col or not price_col:
        return

    start_row = data_row_start
    end_row = data_row_start + row_count - 1
    output_row = end_row + 2
    have_letter = get_column_letter(have_col)
    price_letter = get_column_letter(price_col)

    ws.cell(row=output_row, column=1, value="Total Cost (Have Physical Copy = false)")
    formula = (
        f"=SUMPRODUCT(--(LOWER(TEXT(${have_letter}${start_row}:${have_letter}${end_row},\"@\"))=\"false\"),"
        f"IFERROR(VALUE(SUBSTITUTE(${price_letter}${start_row}:${price_letter}${end_row},\"$\",\"\")),0))"
    )
    total_cell = ws.cell(row=output_row, column=price_col, value=formula)
    total_cell.number_format = "$0.00"


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Convert deck CSV to XLSX for Google Sheets.")
    parser.add_argument("csv_path", help="Path to the deck CSV file")
    parser.add_argument("xlsx_path", help="Output XLSX path")
    args = parser.parse_args()

    csv_path = Path(args.csv_path)
    xlsx_path = Path(args.xlsx_path)

    deck_header, deck_rows = parse_csv(csv_path)
    wb = Workbook()
    deck_ws = wb.active
    deck_ws.title = "Deck"
    write_sheet(deck_ws, deck_header, deck_rows, freeze_row=1, scale=6)
    apply_conditional_formatting(deck_ws, deck_header)
    add_missing_copy_total(deck_ws, deck_header, len(deck_rows))

    formatting_ws = wb.create_sheet("Formatting")
    formatting_ws.cell(row=1, column=1, value="Conditional formatting rules")
    formatting_ws.cell(row=2, column=1, value="Have Physical Copy: TRUE = green, FALSE = yellow")
    formatting_ws.cell(row=3, column=1, value="Cut: TRUE = red")

    header_map = {name: idx + 1 for idx, name in enumerate(deck_header)}
    have_col = header_map.get("Have Physical Copy")
    cut_col = header_map.get("Cut")
    have_letter = get_column_letter(have_col) if have_col else "C"
    cut_letter = get_column_letter(cut_col) if cut_col else "D"

    formatting_ws.cell(row=5, column=1, value="Have Physical Copy formula (green)")
    formatting_ws.cell(
        row=6,
        column=1,
        value=(
            f"=OR(${have_letter}2=TRUE,LOWER(TEXT(${have_letter}2,\"@\"))=\"true\")"
        ),
    )
    formatting_ws.cell(row=7, column=1, value="Have Physical Copy formula (yellow)")
    formatting_ws.cell(
        row=8,
        column=1,
        value=(
            f"=OR(${have_letter}2=FALSE,LOWER(TEXT(${have_letter}2,\"@\"))=\"false\")"
        ),
    )
    formatting_ws.cell(row=9, column=1, value="Cut formula (red)")
    formatting_ws.cell(
        row=10,
        column=1,
        value=f"=OR(${cut_letter}2=TRUE,LOWER(TEXT(${cut_letter}2,\"@\"))=\"true\")",
    )

    wb.save(xlsx_path)


if __name__ == "__main__":
    main()
